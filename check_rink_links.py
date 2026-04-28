"""
RinkLinker — Broken Link Checker
Checks all URLs in rinks_enriched_working.csv and outputs a report.
Run: python check_rink_links.py
Output: link_check_report.xlsx
"""

import pandas as pd
import requests
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings("ignore")

CSV_PATH = "rinks_enriched_working.csv"
OUTPUT_PATH = "link_check_report.xlsx"
TIMEOUT = 8
MAX_WORKERS = 20

URL_COLUMNS = [
    "website",
    "facebook_url",
    "instagram_url",
    "tiktok_url",
    "registration_url",
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

def check_url(url):
    if not url or str(url).strip() in ("", "nan", "None"):
        return "empty", None
    url = str(url).strip()
    if not url.startswith("http"):
        url = "https://" + url
    try:
        r = requests.head(url, timeout=TIMEOUT, headers=HEADERS, allow_redirects=True)
        if r.status_code < 400:
            return "ok", r.status_code
        r2 = requests.get(url, timeout=TIMEOUT, headers=HEADERS, allow_redirects=True)
        if r2.status_code < 400:
            return "ok", r2.status_code
        return "broken", r2.status_code
    except requests.exceptions.ConnectionError:
        return "broken", "connection error"
    except requests.exceptions.Timeout:
        return "timeout", "timed out"
    except Exception as e:
        return "broken", str(e)[:50]

def check_rink(row):
    results = {"id": row.get("id"), "name": row.get("name"), "state": row.get("state")}
    for col in URL_COLUMNS:
        url = row.get(col, "")
        status, code = check_url(url)
        results[f"{col}_url"] = url if url and str(url) != "nan" else ""
        results[f"{col}_status"] = status
        results[f"{col}_code"] = code
    return results

def main():
    print(f"Reading {CSV_PATH}...")
    df = pd.read_csv(CSV_PATH)
    total = len(df)
    print(f"Found {total} rinks. Checking URLs with {MAX_WORKERS} workers...")

    rows = df.to_dict(orient="records")
    results = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(check_rink, row): row for row in rows}
        done = 0
        for future in as_completed(futures):
            results.append(future.result())
            done += 1
            if done % 100 == 0:
                print(f"  {done}/{total} checked...")

    results.sort(key=lambda x: x.get("id", 0))

    wb = Workbook()

    # ── Sheet 1: Broken links only ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Broken Links"

    navy = "0D1B2E"
    red_fill = PatternFill("solid", start_color="FCEBEB")
    yellow_fill = PatternFill("solid", start_color="FAEEDA")
    white = "FFFFFF"

    headers = ["ID", "Name", "State", "Field", "URL", "Status", "Code"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=white, name="Arial")
        cell.fill = PatternFill("solid", start_color=navy)
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    broken_count = 0
    timeout_count = 0

    for r in results:
        for col in URL_COLUMNS:
            status = r.get(f"{col}_status", "empty")
            url = r.get(f"{col}_url", "")
            code = r.get(f"{col}_code", "")
            if status in ("broken", "timeout") and url:
                fill = red_fill if status == "broken" else yellow_fill
                data = [r["id"], r["name"], r["state"], col.replace("_url",""), url, status, code]
                for c, val in enumerate(data, 1):
                    cell = ws1.cell(row=row_num, column=c, value=val)
                    cell.font = Font(name="Arial", size=10)
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True)
                row_num += 1
                if status == "broken":
                    broken_count += 1
                else:
                    timeout_count += 1

    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 50
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 18

    # ── Sheet 2: Full report ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Full Report")
    full_headers = ["ID", "Name", "State"]
    for col in URL_COLUMNS:
        label = col.replace("_url", "").replace("_", " ").title()
        full_headers += [f"{label} URL", f"{label} Status"]

    for c, h in enumerate(full_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=white, name="Arial")
        cell.fill = PatternFill("solid", start_color=navy)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ok_fill = PatternFill("solid", start_color="EAF3DE")

    for i, r in enumerate(results, 2):
        ws2.cell(row=i, column=1, value=r["id"]).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=2, value=r["name"]).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=3, value=r["state"]).font = Font(name="Arial", size=10)
        col_idx = 4
        for col in URL_COLUMNS:
            url = r.get(f"{col}_url", "")
            status = r.get(f"{col}_status", "empty")
            url_cell = ws2.cell(row=i, column=col_idx, value=url)
            url_cell.font = Font(name="Arial", size=10)
            status_cell = ws2.cell(row=i, column=col_idx+1, value=status)
            status_cell.font = Font(name="Arial", size=10)
            if status == "ok":
                status_cell.fill = ok_fill
            elif status == "broken":
                status_cell.fill = red_fill
            elif status == "timeout":
                status_cell.fill = yellow_fill
            col_idx += 2

    for col in range(1, len(full_headers)+1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 20

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ok_total = sum(
        1 for r in results for col in URL_COLUMNS
        if r.get(f"{col}_status") == "ok"
    )
    empty_total = sum(
        1 for r in results for col in URL_COLUMNS
        if r.get(f"{col}_status") == "empty"
    )

    summary = [
        ["The Rink Link — Link Check Report"],
        [""],
        ["Run date", datetime.now().strftime("%B %d, %Y %I:%M %p")],
        ["Total rinks checked", total],
        [""],
        ["RESULTS", "Count"],
        ["OK", ok_total],
        ["Broken", broken_count],
        ["Timeout", timeout_count],
        ["Empty (no URL)", empty_total],
        [""],
        ["BROKEN BY FIELD", "Count"],
    ]

    for col in URL_COLUMNS:
        count = sum(1 for r in results if r.get(f"{col}_status") == "broken")
        label = col.replace("_url", "").replace("_", " ").title()
        summary.append([label, count])

    for row in summary:
        ws3.append(row)

    ws3["A1"].font = Font(bold=True, size=14, name="Arial")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15

    wb.save(OUTPUT_PATH)
    print(f"\nDone!")
    print(f"  Broken links: {broken_count}")
    print(f"  Timeouts:     {timeout_count}")
    print(f"  Report saved: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
